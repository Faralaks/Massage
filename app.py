#!/usr/bin/python3
from flask import Flask, g, redirect, render_template, request, url_for, session, send_file
import sqlite3
from openpyxl import Workbook
from io import  BytesIO
from  datetime import date, timedelta
from config import  DB_PATH, HOST, PORT, LOGIN, PAS, TELEGRAM_TOKEN
from threading import Thread
import bot

path = DB_PATH

week_days = {1:"Понедельник",2:"Вторник",3:"Среда",4:"Четверг",5:"Пятница",6:"Суббота",7:"Воскресенье",}

app = Flask(__name__)
app.config.from_object('config')

h1 = '<h1>{}</h1>'
cap = lambda s: " ".join(list(map(lambda st: st.capitalize(), s.split())))
up = lambda s : s.strip().upper()

form = lambda key: request.form[key]
form_get = lambda key, ret: request.form.get(key, ret)

def p(*items):
    print('\n-----------------------\n')
    for i in items:
        print('\t', i)
    print('\n-----------------------\n')







def get_db():
    db = getattr(g, 'db', None)
    if db is None:
        db = g.db = sqlite3.connect(path)
    return db


@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, 'db', None)
    if db is not None:
        db.close()


@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('login'): return redirect(url_for('index'))
    if request.method == 'POST':

        if request.form.get('login', None).lower() == LOGIN and request.form.get('password', None) == PAS:
            session['login'] = True
            return redirect(url_for('index'))

        return redirect(url_for('login'))

    else:
        return render_template('login.html')


@app.route('/download/<when>')
def download(when='cur'):
    if not session.get('login'): return redirect(url_for('login'))
    stream = BytesIO()
    xlsx = Workbook()
    sheet = xlsx.active
    db = get_db()
    cur = db.cursor()

    if when != 'cur':
        today = date.today()
        first = today.replace(day=1)
        d = first - timedelta(days=1)
    else:
        d = date.today()

    sheet = xlsx['Sheet']
    xlsx.remove(sheet)

    sheet = xlsx.create_sheet('Даты')
    dates = cur.execute('SELECT d, m, y, uid FROM proc WHERE  m=? AND y=? ORDER BY uid', (d.month, d.year)).fetchall()
    date_dict = {}
    for i in dates:
        dt = (i[2], i[1], i[0])
        if date_dict.get(dt) is None: date_dict[dt] = []
        date_dict[dt].append(i[3])
    sorted_dates = sorted(date_dict.items(), key=lambda x: x[0][2])
    for i in sorted_dates:
        sheet.append(['%d.%d.%d' % (i[0][2], i[0][1], i[0][0]), len(i[1])]+i[1])



    sheet = xlsx.create_sheet('Процедуры')
    procs = cur.execute('SELECT * FROM proc WHERE  m=? AND y=? ORDER BY uid', (d.month, d.year)).fetchall()
    sheet.append(['ФИО', 'Класс', 'Сделано процедур'])
    proc = {}
    for i in procs:
        if proc.get(i[0]) is None: proc[i[0]] = []
        proc[i[0]].append('%d.%d.%d'%(i[1], i[2], i[3]))


    for i in proc.items():
        fam, grade = i[0].split(' — ')[:2]
        fam = cap(fam)
        sheet.append([fam, grade, len(i[1])]+sorted(i[1], key=lambda x: int(x.split('.')[0])))


    xlsx.save(stream)
    xlsx.close()
    stream.seek(0)
    return send_file(stream, cache_timeout=0, as_attachment=True, attachment_filename='massage_counter.xlsx')






@app.route('/addNew', methods=['POST'])
def addNew():
    if not session.get('login'): return redirect(url_for('login'))
    if form_get('fam', None) is None or form_get('grade', None) is None or form_get('fam', None) == '' or form_get('grade', None) == '': return h1.format('Не получено имя или класс ученика')
    db = get_db()
    cur = db.cursor()
    cur.execute('INSERT INTO people VALUES (?,0,?)', (cap(form('fam'))+' — '+up(form('grade')),None))
    db.commit()
    return redirect(url_for('index'))

@app.route('/add', methods=['POST'])
def add():
    if not session.get('login'): return redirect(url_for('login'))
    if form_get('list', None) is None or form_get('list', None) == 'Выберите ученика': return h1.format('Не был выбран ученик')
    if form_get('date', None) is None: return h1.format('Не была получена дата процедуры')
    db = get_db()
    cur = db.cursor()
    form_data = form('list').split(' — ')
    proc_date = form('date').split(' — ')[1].split('.')
    uid = form_data[0] +' — '+ form_data[1]
    data = cur.execute("SELECT count FROM people WHERE uid=?", (uid, )).fetchone()
    cur.execute('UPDATE people SET count=?, last=? WHERE uid=?', [data[0]+1, ".".join(proc_date), uid])
    cur.execute('INSERT INTO proc VALUES (?,?,?,?)', (uid, proc_date[0], proc_date[1], proc_date[2]))
    db.commit()
    return redirect(url_for('index'))


@app.route('/delete', methods=['POST'])
def delete():
    if not session.get('login'): return redirect(url_for('login'))
    if form_get('delList', None) is None or form_get('delList', None) == 'Выберите ученика для удаления': return h1.format('Не был выбран ученик для удаления')
    db = get_db()
    cur = db.cursor()
    form_data = form('delList').split(' — ')
    uid = form_data[0] +' — '+ form_data[1]
    data = cur.execute("DELETE FROM people WHERE uid=?", (uid, ))
    db.commit()
    return redirect(url_for('index'))




@app.route('/')
def index():
    if not session.get('login'): return redirect(url_for('login'))
    db = get_db()
    cur = db.cursor()
    people = cur.execute('SELECT * FROM people').fetchall()
    today = date.today()
    return render_template('index.html', people=people, str=str, today=today, timedelta=timedelta, week_days=week_days)



@app.route('/change', methods=['POST'])
def change():
    if not session.get('login'): return redirect(url_for('login'))
    if form_get('change', None) is None or form_get('change', None) == 'Выберите ученика для редактирования': return h1.format('Не был выбран ученик для редактирования')
    if form_get('fam', None) is None or form_get('grade', None) is None or form_get('fam', None) == '' or form_get('grade', None) == '': return h1.format('Не получено новое ФИО или Класс ученика')

    db = get_db()
    cur = db.cursor()

    old_uid = form('change')
    new_uid = cap(form("fam"))+' — '+up(form("grade"))
    if old_uid == new_uid: return h1.format("Новые данные равны старым")

    new_count = cur.execute("SELECT count, last FROM people WHERE uid=?", (old_uid, )).fetchone()
    cur_count = cur.execute("SELECT count FROM people WHERE uid=?", (new_uid, )).fetchone()
    if cur_count:
        cur_count = cur_count[0]
    else:
        cur_count = 0
        cur.execute('INSERT INTO people VALUES (?,0,?)', (new_uid, new_count[1]))

    cur.execute('UPDATE people SET count=? WHERE uid=?', [cur_count+new_count[0], new_uid])
    cur.execute('UPDATE proc SET uid=? WHERE uid=?', [new_uid, old_uid])
    cur.execute("DELETE FROM people WHERE uid=?", (old_uid, ))

    db.commit()
    return redirect(url_for('index'))




if __name__ == '__main__':
    tele_bot = Thread(target=bot.run, daemon=True)
    tele_bot.start()
    app.run(host=HOST, port=PORT)
    tele_bot.join()
